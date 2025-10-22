import os
import telebot
from telebot.types import Message, InlineKeyboardMarkup, InlineKeyboardButton
import time
from datetime import datetime, timedelta, timezone
from flask import Flask, request, abort
import threading
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
 
TOKEN = os.environ.get('BOT_TOKEN')
BOT_URL = '/webhook'
ADMIN_CHAT_ID = os.environ.get('ADMIN_CHAT_ID')

if not TOKEN or not ADMIN_CHAT_ID:
    raise ValueError("Не установлены BOT_TOKEN или ADMIN_CHAT_ID")

bot = telebot.TeleBot(TOKEN)

positions = {
    'Ватрушка': 200, 'Капуста': 130, 'Яблоко': 120, 'Картофель': 130,
    'Мак': 190, 'Плюшка': 150, 'Чечевица': 140, 'Повидло': 130,
    'Корица': 150, 'Сосиск в тесте': 150, 'Брусника': 130,
    'Вишня': 130, 'Черная смородина': 130, 'Творог с зеленью': 130
}

# Файлы для хранения данных
DATA_DIR = '/data'
if not os.path.exists(DATA_DIR):
    os.makedirs(DATA_DIR)

USERS_DB_FILE = os.path.join(DATA_DIR, 'users_data.json')
ORDERS_DB_FILE = os.path.join(DATA_DIR, 'orders_history.json')
STATE_FILE = os.path.join(DATA_DIR, 'scheduler_state.json')

# Временные данные (в оперативной памяти)
current_orders = {}
registration_steps = {}

app = Flask(__name__)

# === ФУНКЦИИ РАБОТЫ С ДАННЫМИ ===

def load_users_data():
    """Загрузка базы клиентов из файла"""
    try:
        if os.path.exists(USERS_DB_FILE):
            with open(USERS_DB_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        print(f"Ошибка загрузки users_data: {e}")
    return {}

def save_users_data():
    """Сохранение базы клиентов в файл"""
    try:
        with open(USERS_DB_FILE, 'w', encoding='utf-8') as f:
            json.dump(users_data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Ошибка сохранения users_data: {e}")

def load_orders_history():
    """Загрузка истории заказов из файла"""
    try:
        if os.path.exists(ORDERS_DB_FILE):
            with open(ORDERS_DB_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        print(f"Ошибка загрузки orders_history: {e}")
    return {}

def save_orders_history():
    """Сохранение истории заказов в файл"""
    try:
        with open(ORDERS_DB_FILE, 'w', encoding='utf-8') as f:
            json.dump(orders_history, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Ошибка сохранения orders_history: {e}")

def add_order_to_history(user_data, date_str):
    """Добавление заказа в историю"""
    try:
        if date_str not in orders_history:
            orders_history[date_str] = []
        
        order_entry = {
            'user_id': user_data.get('user_id'),
            'location_name': user_data['location_name'],
            'address': user_data['address'],
            'orders': user_data['orders'].copy(),
            'total_items': sum(user_data['orders'].values()),
            'timestamp': datetime.now().strftime('%H:%M')
        }
        
        orders_history[date_str].append(order_entry)
        save_orders_history()
    except Exception as e:
        print(f"Ошибка добавления в историю: {e}")

def load_scheduler_state():
    """Загрузка состояния планировщика"""
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"Ошибка загрузки состояния: {e}")
    return {"target_send_minute": None, "target_clear_minute": None, "last_triggered_minute": None}

def save_scheduler_state(state):
    """Сохранение состояния планировщика"""
    try:
        with open(STATE_FILE, "w", encoding="utf-8") as f:
            json.dump(state, f)
    except Exception as e:
        print(f"Ошибка сохранения состояния: {e}")

# Загрузка данных при запуске
users_data = load_users_data()
orders_history = load_orders_history()

print(f"Загружено пользователей: {len(users_data)}")
print(f"Загружено дней в истории: {len(orders_history)}")

# === FLASK WEBHOOK ===

@app.route(BOT_URL, methods=['POST'])
def webhook():
    print(f"ПОЛУЧЕН POST на {BOT_URL}")
    try:
        if request.headers.get('content-type') == 'application/json':
            json_string = request.get_data().decode('utf-8')
            update = telebot.types.Update.de_json(json_string)
            bot.process_new_updates([update])
            return '', 200
        else:
            print("ОТКЛОНЁН: не JSON")
            abort(403)
    except Exception as e:
        print(f"ОШИБКА В WEBHOOK: {e}")
        return 'Error', 500

@app.route('/')
def index():
    return "Бот работает на Railway!"

# === ФУНКЦИИ БОТА ===

def get_user_data(user_id):
    """Получить данные пользователя"""
    user_id_str = str(user_id)
    if user_id_str not in users_data:
        users_data[user_id_str] = {
            'user_id': user_id_str,
            'address': '',
            'location_name': '', 
            'orders': {},
            'registered': False,
            'registration_date': datetime.now().strftime('%d.%m.%Y %H:%M')
        }
        save_users_data()
    return users_data[user_id_str]

@bot.message_handler(commands=['start'])
def start(message: Message):
    user_id = message.from_user.id
    user_data = get_user_data(user_id)
    
    if not user_data['registered']:
        start_registration(message)
    else:
        show_main_menu(message.chat.id, user_data)

def start_registration(message):
    """Начать процесс регистрации"""
    user_id = message.from_user.id
    registration_steps[user_id] = 'waiting_location'
    
    bot.send_message(
        message.chat.id,
        "Добро пожаловать! Давайте зарегистрируем вас.\n\n"
        "**Как называется ваша точка/магазин?**\n"
        "Например: 'Магазин у дома', 'Офис на Ленина', 'Кафе Уют'"
    )

@bot.message_handler(commands=['admin'])
def admin_panel(message: Message):
    """Панель администратора"""
    if str(message.chat.id) != ADMIN_CHAT_ID:
        bot.reply_to(message, "Доступ запрещен")
        return
    
    markup = InlineKeyboardMarkup(row_width=2)
    buttons = [
        InlineKeyboardButton('Excel Сводка', callback_data='admin_excel'),
        InlineKeyboardButton('Текстовая сводка', callback_data='admin_summary'),
        InlineKeyboardButton('База клиентов', callback_data='admin_clients'),
        InlineKeyboardButton('Удалить клиентов', callback_data='admin_delete_clients'),
        InlineKeyboardButton('История заказов', callback_data='admin_history'),
        InlineKeyboardButton('Обнулить заказы', callback_data='admin_clear'),
        InlineKeyboardButton('Экспорт данных', callback_data='admin_export'),
    ]
    markup.add(*buttons)
    
    stats_text = f"**Статистика:**\nКлиентов: {len(users_data)}\nДней в истории: {len(orders_history)}"
    
    bot.send_message(message.chat.id, f"**Панель администратора**\n\n{stats_text}", reply_markup=markup)

@bot.message_handler(func=lambda message: True)
def handle_messages(message: Message):
    user_id = message.from_user.id
    
    # Обработка регистрации
    if user_id in registration_steps:
        handle_registration(message)
        return
    
    # Обработка количества для заказов
    if user_id in current_orders:
        handle_quantity(message)
        return
    
    bot.reply_to(message, "Используйте меню для навигации")

def handle_registration(message: Message):
    """Обработка шагов регистрации"""
    user_id = message.from_user.id
    step = registration_steps.get(user_id)
    user_data = get_user_data(user_id)
    
    if step == 'waiting_location':
        user_data['location_name'] = message.text.strip()
        registration_steps[user_id] = 'waiting_address'
        
        bot.send_message(
            message.chat.id,
            "**Теперь укажите адрес доставки:**\n"
            "Например: 'ул. Ленина, 15', 'ТЦ Центральный, 2 этаж'"
        )
        
    elif step == 'waiting_address':
        user_data['address'] = message.text.strip()
        user_data['registered'] = True
        user_data['registration_date'] = datetime.now().strftime('%d.%m.%Y %H:%M')
        del registration_steps[user_id]
        
        save_users_data()
        
        bot.send_message(
            message.chat.id,
            f"**Регистрация завершена!**\n\n"
            f"Точка: {user_data['location_name']}\n"
            f"Адрес: {user_data['address']}\n\n"
            f"Теперь вы можете делать заказы!"
        )
        
        show_main_menu(message.chat.id, user_data)

def show_main_menu(chat_id, user_data):
    """Показать главное меню"""
    markup = InlineKeyboardMarkup(row_width=2)
    buttons = [
        InlineKeyboardButton('Добавить заказ', callback_data='add_order'),
        InlineKeyboardButton('Мой заказ', callback_data='my_order'),
        InlineKeyboardButton('Изменить заказ', callback_data='edit_order'),
        InlineKeyboardButton('Мои данные', callback_data='my_data'),
    ]
    markup.add(*buttons)
    
    welcome_text = f"{user_data['location_name']}\n{user_data['address']}\n\nВыберите действие:"
    bot.send_message(chat_id, welcome_text, reply_markup=markup)

@bot.callback_query_handler(func=lambda call: True)
def handle_callback(call):
    user_id = call.from_user.id
    chat_id = call.message.chat.id
    user_data = get_user_data(user_id)
    
    if call.data == 'add_order':
        show_positions_menu(chat_id)
    elif call.data == 'my_order':
        show_user_order(call, user_data)
    elif call.data == 'edit_order':
        show_edit_menu(call, user_data)
    elif call.data == 'my_data':
        show_user_data(call, user_data)
    elif call.data == 'admin_excel':
        send_excel_summary(call)
    elif call.data == 'admin_summary':
        send_text_summary(call)
    elif call.data == 'admin_clients':
        show_clients_database(call)
    elif call.data == 'admin_delete_clients':
        show_delete_clients_menu(call)
    elif call.data == 'admin_history':
        show_orders_history(call)
    elif call.data == 'admin_clear':
        clear_all_orders(call)
    elif call.data == 'admin_export':
        export_all_data(call)
    elif call.data in positions:
        current_orders[user_id] = {'position': call.data}
        bot.answer_callback_query(call.id, f"Выбрано: {call.data}")
        bot.send_message(chat_id, f"Сколько штук {call.data}?")
    elif call.data.startswith('edit_'):
        position = call.data[5:]
        current_orders[user_id] = {'position': position, 'editing': True}
        bot.answer_callback_query(call.id, f"Изменяем: {position}")
        bot.send_message(chat_id, f"Введите новое количество для {position}:")
    elif call.data == 'back_to_main':
        bot.answer_callback_query(call.id, "Возврат в меню")
        bot.delete_message(chat_id, call.message.message_id)
        show_main_menu(chat_id, user_data)
    elif call.data == 'clear_order':
        user_data['orders'] = {}
        save_users_data()
        bot.answer_callback_query(call.id, "Заказ очищен")
        bot.delete_message(chat_id, call.message.message_id)
        show_main_menu(chat_id, user_data)
    elif call.data.startswith('delete_user_'):
        delete_user(call)
    elif call.data == 'admin_stats':
        bot.answer_callback_query(call.id, "Детальная статистика")
        bot.send_message(chat_id, "Детальная статистика (в разработке)")
    elif call.data == 'back_to_admin':
        bot.answer_callback_query(call.id)
        bot.delete_message(chat_id, call.message.message_id)
        admin_panel(call.message)

def show_positions_menu(chat_id):
    markup = InlineKeyboardMarkup(row_width=2)
    for pos in positions.keys():
        markup.add(InlineKeyboardButton(pos, callback_data=pos))
    markup.add(InlineKeyboardButton('Назад', callback_data='back_to_main'))
    
    bot.send_message(chat_id, "Выберите позицию для заказа:", reply_markup=markup)

def show_user_order(call, user_data):
    user_orders = user_data['orders']
    
    if not user_orders:
        bot.answer_callback_query(call.id, "У вас нет заказов")
        bot.send_message(call.message.chat.id, "У вас еще нет заказов на сегодня.")
        return
    
    total_items = sum(user_orders.values())
    
    order_text = f"**{user_data['location_name']}**\n"
    order_text += f"{user_data['address']}\n\n"
    order_text += "**Ваш заказ на сегодня:**\n\n"
    
    for pos, qty in user_orders.items():
        order_text += f"• {pos}: {qty} шт.\n"
    
    order_text += f"\n**Итого:** {total_items} шт."
    
    bot.answer_callback_query(call.id)
    bot.send_message(call.message.chat.id, order_text)

def show_user_data(call, user_data):
    user_orders = user_data['orders']
    total_items = sum(user_orders.values()) if user_orders else 0
    
    data_text = "**Ваши данные:**\n\n"
    data_text += f"**Точка:** {user_data['location_name']}\n"
    data_text += f"**Адрес:** {user_data['address']}\n"
    data_text += f"**Дата регистрации:** {user_data.get('registration_date', 'неизвестно')}\n"
    data_text += f"**Заказов сегодня:** {total_items} шт.\n\n"
    data_text += "_Чтобы изменить данные, перезапустите бота /start_"
    
    bot.answer_callback_query(call.id)
    bot.send_message(call.message.chat.id, data_text)

def show_edit_menu(call, user_data):
    user_orders = user_data['orders']
    
    if not user_orders:
        bot.answer_callback_query(call.id, "Нет заказов для редактирования")
        return
    
    markup = InlineKeyboardMarkup(row_width=2)
    
    for pos in user_orders.keys():
        markup.add(InlineKeyboardButton(f"{pos}", callback_data=f'edit_{pos}'))
    
    markup.add(InlineKeyboardButton('Добавить еще', callback_data='add_order'))
    markup.add(InlineKeyboardButton('Очистить все', callback_data='clear_order'))
    markup.add(InlineKeyboardButton('Назад', callback_data='back_to_main'))
    
    bot.answer_callback_query(call.id)
    bot.send_message(call.message.chat.id, "Выберите позицию для изменения:", reply_markup=markup)

def handle_quantity(message: Message):
    user_id = message.from_user.id
    chat_id = message.chat.id
    user_data = get_user_data(user_id)
    
    position_data = current_orders[user_id]
    position = position_data['position']
    is_editing = position_data.get('editing', False)
    
    try:
        quantity = int(message.text.strip())
        if quantity < 0:
            raise ValueError
        
        if quantity == 0:
            if position in user_data['orders']:
                del user_data['orders'][position]
            action_text = f"Удалено: {position}"
        else:
            user_data['orders'][position] = quantity
            action_text = f"{'Обновлено' if is_editing else 'Добавлено'} {quantity} шт. {position}"
        
        save_users_data()
        
        bot.reply_to(message, f"{action_text} для {user_data['location_name']}!")
        del current_orders[user_id]
        
        show_main_menu(chat_id, user_data)
        
    except ValueError:
        bot.reply_to(message, "Введите целое число (0 для удаления позиции):")

# === ГЕНЕРАЦИЯ EXCEL ===

def generate_excel_file():
    """Генерация Excel файла со сводкой"""
    active_users = [data for data in users_data.values() if data.get('orders') and data.get('registered')]
    
    if not active_users:
        return None
    
    active_users.sort(key=lambda x: x['location_name'])
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка заказов"
    
    # Стили
    header_font = Font(bold=True, size=14)
    title_font = Font(bold=True, size=12)
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Заголовок
    num_positions = len(positions)
    header_end_col = get_column_letter(3 + num_positions + 1)
    ws.merge_cells(f'A1:{header_end_col}1')
    ws['A1'] = f"Сводка заказов от {datetime.now().strftime('%d.%m.%Y')}"
    ws['A1'].font = header_font
    ws['A1'].alignment = center_align
    
    ws.append([])
    
    headers = ['№', 'Точка', 'Адрес'] + list(positions.keys()) + ['ИТОГО']
    ws.append(headers)
    
    header_row = 3
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = title_font
        cell.alignment = center_align
        cell.border = border
    
    row_num = 4
    for i, user_data in enumerate(active_users, 1):
        row = [i, user_data['location_name'], user_data['address']]
        total = 0
        
        for pos in positions.keys():
            qty = user_data['orders'].get(pos, 0)
            row.append(qty)
            total += qty
        
        row.append(total)
        ws.append(row)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border
            if col in [1, len(headers)]:
                cell.font = bold_font
        row_num += 1
    
    ws.append([])
    row_num += 1
    total_row = ['ВСЕГО', '', '']
    
    for pos_idx in range(len(positions)):
        col_idx = 4 + pos_idx
        pos_total = sum(ws.cell(row=r, column=col_idx).value or 0 for r in range(4, row_num))
        total_row.append(pos_total)
    
    grand_total = sum(total_row[3:])
    total_row.append(grand_total)
    ws.append(total_row)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row_num + 1, column=col)
        cell.font = bold_font
        cell.border = border
        if col >= 4:
            cell.alignment = center_align
    
    column_widths = {'A': 5, 'B': 25, 'C': 30}
    for i, pos in enumerate(positions.keys(), 4):
        col_letter = get_column_letter(i)
        column_widths[col_letter] = 8
    column_widths[get_column_letter(len(headers))] = 10
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

def send_excel_summary(call=None):
    """Отправка Excel сводки"""
    try:
        excel_buffer = generate_excel_file()
        
        if not excel_buffer:
            message = "Нет заказов за сегодня."
            if call:
                bot.answer_callback_query(call.id, "Нет заказов")
                bot.send_message(call.message.chat.id, message)
            else:
                bot.send_message(ADMIN_CHAT_ID, message)
            return
        
        filename = f"заказы_{datetime.now().strftime('%d.%m.%Y')}.xlsx"
        excel_buffer.name = filename
        
        input_file = telebot.types.InputFile(excel_buffer)
        
        if call:
            bot.answer_callback_query(call.id)
            bot.send_document(
                call.message.chat.id,
                document=input_file,
                caption=f"Сводка заказов от {datetime.now().strftime('%d.%m.%Y')}"
            )
        else:
            # Сохраняем в историю
            current_date = datetime.now().strftime('%Y-%m-%d')
            for user_data in users_data.values():
                if user_data.get('orders') and user_data.get('registered'):
                    add_order_to_history(user_data, current_date)
            
            bot.send_document(
                ADMIN_CHAT_ID,
                document=input_file,
                caption=f"Автоматическая сводка заказов от {datetime.now().strftime('%d.%m.%Y')}"
            )
            
    except Exception as e:
        print(f"Ошибка при отправке сводки: {e}")
        error_message = f"Ошибка при генерации Excel: {e}"
        if call:
            bot.send_message(call.message.chat.id, error_message)
        else:
            bot.send_message(ADMIN_CHAT_ID, error_message)

# === АДМИНИСТРАТИВНЫЕ ФУНКЦИИ ===

def send_text_summary(call):
    """Текстовая сводка"""
    active_users = [data for data in users_data.values() if data.get('orders') and data.get('registered')]
    
    if not active_users:
        bot.answer_callback_query(call.id, "Нет заказов")
        bot.send_message(call.message.chat.id, "Нет заказов за сегодня.")
        return
    
    active_users.sort(key=lambda x: x['location_name'])
    
    summary_text = f"**Сводка заказов от {datetime.now().strftime('%d.%m.%Y')}**\n"
    summary_text += f"Клиентов: {len(active_users)}\n\n"
    
    for user_data in active_users:
        total_items = sum(user_data['orders'].values())
        order_details = [f"{pos}:{qty}" for pos, qty in user_data['orders'].items() if qty > 0]
        
        details_str = ", ".join(order_details)
        summary_text += f"• **{user_data['location_name']}** - {total_items} шт.\n"
        summary_text += f"  {details_str}\n"
        summary_text += f"  {user_data['address']}\n\n"
    
    bot.answer_callback_query(call.id)
    bot.send_message(call.message.chat.id, summary_text)

def show_clients_database(call):
    """Показать базу клиентов"""
    registered_users = [data for data in users_data.values() if data['registered']]
    
    if not registered_users:
        bot.answer_callback_query(call.id, "База клиентов пуста")
        bot.send_message(call.message.chat.id, "База клиентов пуста.")
        return
    
    clients_text = f"**БАЗА КЛИЕНТОВ**\nВсего: {len(registered_users)}\n\n"
    
    for i, user_data in enumerate(registered_users, 1):
        order_count = sum(user_data['orders'].values())
        last_order = "Сегодня" if order_count > 0 else "Нет заказов"
        clients_text += f"{i}. **{user_data['location_name']}**\n"
        clients_text += f"   {user_data['address']}\n"
        clients_text += f"   Регистрация: {user_data.get('registration_date', 'неизвестно')}\n"
        clients_text += f"   {last_order} ({order_count} шт.)\n\n"
    
    bot.answer_callback_query(call.id)
    bot.send_message(call.message.chat.id, clients_text)

def show_delete_clients_menu(call):
    if str(call.message.chat.id) != ADMIN_CHAT_ID:
        bot.answer_callback_query(call.id, "Доступ запрещен")
        return
    
    registered_users = [data for data in users_data.values() if data['registered']]
    
    if not registered_users:
        bot.answer_callback_query(call.id, "Нет клиентов")
        bot.send_message(call.message.chat.id, "Нет клиентов для удаления.")
        return
    
    markup = InlineKeyboardMarkup(row_width=1)
    for user_data in registered_users:
        button_text = f"Удалить {user_data['location_name']}"
        markup.add(InlineKeyboardButton(button_text, callback_data=f"delete_user_{user_data['user_id']}"))
    
    markup.add(InlineKeyboardButton('Назад в админ', callback_data='back_to_admin'))
    
    bot.answer_callback_query(call.id)
    bot.send_message(call.message.chat.id, "Выберите клиента для удаления:", reply_markup=markup)

def delete_user(call):
    if str(call.message.chat.id) != ADMIN_CHAT_ID:
        bot.answer_callback_query(call.id, "Доступ запрещен")
        return
    
    user_id_str = call.data.split('_')[-1]
    if user_id_str in users_data:
        location_name = users_data[user_id_str]['location_name']
        del users_data[user_id_str]
        save_users_data()
        bot.answer_callback_query(call.id, f"Клиент {location_name} удален")
        bot.delete_message(call.message.chat.id, call.message.message_id)
        bot.send_message(call.message.chat.id, f"Клиент {location_name} удален из базы.")
    else:
        bot.answer_callback_query(call.id, "Клиент не найден")

def show_orders_history(call):
    """Показать историю заказов"""
    if not orders_history:
        bot.answer_callback_query(call.id, "История пуста")
        bot.send_message(call.message.chat.id, "История заказов пуста.")
        return
    
    history_text = f"**ИСТОРИЯ ЗАКАЗОВ**\nДней в истории: {len(orders_history)}\n\n"
    
    sorted_dates = sorted(orders_history.keys(), reverse=True)[:7]
    
    for date_str in sorted_dates:
        date_orders = orders_history[date_str]
        total_orders = len(date_orders)
        total_items = sum(order['total_items'] for order in date_orders)
        
        history_text += f"**{datetime.strptime(date_str, '%Y-%m-%d').strftime('%d.%m.%Y')}**\n"
        history_text += f"   Клиентов: {total_orders}\n"
        history_text += f"   Товаров: {total_items} шт.\n\n"
    
    markup = InlineKeyboardMarkup()
    markup.add(InlineKeyboardButton('Детальная статистика', callback_data='admin_stats'))
    
    bot.answer_callback_query(call.id)
    bot.send_message(call.message.chat.id, history_text, reply_markup=markup)

def clear_all_orders(call):
    """Очистить все заказы"""
    cleared_count = 0
    for user_data in users_data.values():
        if user_data['orders']:
            user_data['orders'] = {}
            cleared_count += 1
    
    save_users_data()
    
    bot.answer_callback_query(call.id, f"Очищено {cleared_count}")
    bot.send_message(call.message.chat.id, f"Очищены заказы у {cleared_count} клиентов!")

def clear_all_orders_auto():
    """Автоматическая очистка заказов"""
    cleared_count = 0
    for user_data in users_data.values():
        if user_data['orders']:
            user_data['orders'] = {}
            cleared_count += 1
    
    if cleared_count > 0:
        save_users_data()
    
    print(f"Автоматически очищены заказы у {cleared_count} пользователей")
    return cleared_count

def export_all_data(call):
    """Экспорт всех данных в JSON"""
    try:
        export_data = {
            'users': users_data,
            'orders_history': orders_history,
            'export_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        export_json = json.dumps(export_data, ensure_ascii=False, indent=2).encode('utf-8')
        json_buffer = io.BytesIO(export_json)
        filename = f"backup_data_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
        json_buffer.name = filename
        
        input_file = telebot.types.InputFile(json_buffer)
        
        bot.answer_callback_query(call.id)
        bot.send_document(
            call.message.chat.id,
            document=input_file,
            caption="Полный бэкап данных системы"
        )
    except Exception as e:
        bot.answer_callback_query(call.id, "Ошибка экспорта")
        bot.send_message(call.message.chat.id, f"Ошибка при экспорте данных: {e}")

# === ПЛАНИРОВЩИК ЗАДАЧ ===

# Настройки расписания (МСК)
SCHEDULE_SEND_SUMMARY_TIME = "11:03"  # Время отправки сводки
SCHEDULE_CLEAR_ORDERS_TIME = "11:04"  # Время очистки заказов

def check_scheduled_tasks():
    """Проверка и выполнение запланированных задач"""
    state = load_scheduler_state()
    msk_tz = timezone(timedelta(hours=3))
    now = datetime.now(msk_tz)
    current_time = now.strftime('%H:%M')
    current_date = now.strftime('%Y-%m-%d')
    
    print(f"--- ПРОВЕРКА: {current_time}:{now.strftime('%S')} МСК ---")

    # Инициализируем состояние при первом запуске
    if "last_send_date" not in state:
        state["last_send_date"] = None
    if "last_clear_date" not in state:
        state["last_clear_date"] = None
    
    # === ОТПРАВКА СВОДКИ ===
    if current_time == SCHEDULE_SEND_SUMMARY_TIME:
        # Проверяем, что сегодня ещё не отправляли
        if state["last_send_date"] != current_date:
            print("*** ТРИГГЕР: ОТПРАВКА СВОДКИ ***")
            try:
                send_excel_summary()
                state["last_send_date"] = current_date
                save_scheduler_state(state)
                print(f"✅ Сводка отправлена в {current_time}!")
            except Exception as e:
                print(f"❌ Ошибка сводки: {e}")
                try:
                    bot.send_message(ADMIN_CHAT_ID, f"❌ Ошибка отправки сводки: {e}")
                except:
                    pass
        else:
            print(f"⏸ Сводка уже отправлена сегодня ({current_date})")
    
    # === ОЧИСТКА ЗАКАЗОВ ===
    if current_time == SCHEDULE_CLEAR_ORDERS_TIME:
        # Проверяем, что сегодня ещё не очищали
        if state["last_clear_date"] != current_date:
            print("*** ТРИГГЕР: ОЧИСТКА ЗАКАЗОВ ***")
            try:
                cleared_count = clear_all_orders_auto()
                state["last_clear_date"] = current_date
                save_scheduler_state(state)
                bot.send_message(ADMIN_CHAT_ID, f"✅ Заказы обнулены в {current_time}. Очищено: {cleared_count}")
                print(f"✅ Очищено: {cleared_count}")
            except Exception as e:
                print(f"❌ Ошибка очистки: {e}")
                try:
                    bot.send_message(ADMIN_CHAT_ID, f"❌ Ошибка очистки заказов: {e}")
                except:
                    pass
        else:
            print(f"⏸ Заказы уже очищены сегодня ({current_date})")
    
    # Логирование раз в 10 проверок (каждые 5 минут при интервале 30 сек)
    check_count = state.get("check_count", 0) + 1
    state["check_count"] = check_count
    
    if check_count % 10 == 0:
        print(f"💤 Ожидаем: сводка в {SCHEDULE_SEND_SUMMARY_TIME}, очистка в {SCHEDULE_CLEAR_ORDERS_TIME} МСК")
        print(f"   Последняя сводка: {state.get('last_send_date', 'никогда')}")
        print(f"   Последняя очистка: {state.get('last_clear_date', 'никогда')}")
        save_scheduler_state(state)

def scheduler():
    """Основной цикл планировщика"""
    print("🚀 ПЛАНИРОВЩИК ЗАПУЩЕН!")
    
    while True:
        try:
            check_scheduled_tasks()
            time.sleep(30)  # Проверяем каждые 30 секунд
        except Exception as e:
            print(f"❌ КРИТИЧЕСКАЯ ОШИБКА ПЛАНИРОВЩИКА: {e}")
            time.sleep(30)

# === ИНИЦИАЛИЗАЦИЯ БОТА ===

def setup_webhook():
    """Установка webhook"""
    print("Удаляю старый webhook...")
    bot.remove_webhook()
    time.sleep(2)

    webhook_url = "https://web-production-d7a9d.up.railway.app/webhook"
    print(f"Устанавливаю webhook: {webhook_url}")

    result = bot.set_webhook(url=webhook_url)
    if result:
        print("✅ WEBHOOK УСПЕШНО УСТАНОВЛЕН!")
    else:
        print("❌ ОШИБКА: Webhook НЕ установлен!")

def start_bot():
    """Запуск бота и планировщика"""
    print("=== ИНИЦИАЛИЗАЦИЯ БОТА ===")
    
    # Установка webhook
    setup_webhook()
    
    # Запуск планировщика в отдельном потоке
    scheduler_thread = threading.Thread(target=scheduler, daemon=True)
    scheduler_thread.start()
    print("✅ ПЛАНИРОВЩИК ЗАПУЩЕН В ОТДЕЛЬНОМ ПОТОКЕ!")
    
    print("✅ БОТ ГОТОВ К РАБОТЕ!")

# Запуск при импорте модуля (безопасно для Gunicorn)
start_bot()
